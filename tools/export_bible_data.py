from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "bible-data.js"

OLD_TESTAMENT = [
    "창세기",
    "출애굽기",
    "레위기",
    "민수기",
    "신명기",
    "여호수아",
    "사사기",
    "룻기",
    "사무엘상",
    "사무엘하",
    "열왕기상",
    "열왕기하",
    "역대상",
    "역대하",
    "에스라",
    "느헤미야",
    "에스더",
    "욥기",
    "시편",
    "잠언",
    "전도서",
    "아가",
    "이사야",
    "예레미야",
    "예레미야애가",
    "에스겔",
    "다니엘",
    "호세아",
    "요엘",
    "아모스",
    "오바댜",
    "요나",
    "미가",
    "나훔",
    "하박국",
    "스바냐",
    "학개",
    "스가랴",
    "말라기",
]

# Book name -> 대한성서공회(bskorea.or.kr) NKRV URL code, e.g.
# https://bible.bskorea.or.kr/bible/NKRV/GEN.1
BOOK_CODES = {
    "창세기": "GEN", "출애굽기": "EXO", "레위기": "LEV", "민수기": "NUM", "신명기": "DEU",
    "여호수아": "JOS", "사사기": "JDG", "룻기": "RUT", "사무엘상": "1SA", "사무엘하": "2SA",
    "열왕기상": "1KI", "열왕기하": "2KI", "역대상": "1CH", "역대하": "2CH", "에스라": "EZR",
    "느헤미야": "NEH", "에스더": "EST", "욥기": "JOB", "시편": "PSA", "잠언": "PRO",
    "전도서": "ECC", "아가": "SNG", "이사야": "ISA", "예레미야": "JER", "예레미야애가": "LAM",
    "에스겔": "EZK", "다니엘": "DAN", "호세아": "HOS", "요엘": "JOL", "아모스": "AMO",
    "오바댜": "OBA", "요나": "JON", "미가": "MIC", "나훔": "NAM", "하박국": "HAB",
    "스바냐": "ZEP", "학개": "HAG", "스가랴": "ZEC", "말라기": "MAL",
    "마태복음": "MAT", "마가복음": "MRK", "누가복음": "LUK", "요한복음": "JHN", "사도행전": "ACT",
    "로마서": "ROM", "고린도전서": "1CO", "고린도후서": "2CO", "갈라디아서": "GAL", "에베소서": "EPH",
    "빌립보서": "PHP", "골로새서": "COL", "데살로니가전서": "1TH", "데살로니가후서": "2TH",
    "디모데전서": "1TI", "디모데후서": "2TI", "디도서": "TIT", "빌레몬서": "PHM", "히브리서": "HEB",
    "야고보서": "JAS", "베드로전서": "1PE", "베드로후서": "2PE", "요한일서": "1JN", "요한이서": "2JN",
    "요한삼서": "3JN", "유다서": "JUD", "요한계시록": "REV",
}


def value(row: tuple, index: int) -> str:
    cell = row[index] if index < len(row) else ""
    return "" if cell is None else str(cell).strip()


def main() -> None:
    candidates = [
        path
        for path in ROOT.glob("*.xlsx")
        if not path.name.startswith("~$") and path.name != "source.xlsx"
    ]
    source = candidates[0] if candidates else ROOT / "source.xlsx"

    if not source.exists():
        raise FileNotFoundError("No .xlsx source workbook found in the project root.")

    try:
        workbook = openpyxl.load_workbook(source, data_only=True)
    except PermissionError:
        source = ROOT / "source.xlsx"
        if not source.exists():
            raise
        workbook = openpyxl.load_workbook(source, data_only=True)
    quiz_sheet = workbook["전체 퀴즈"]

    chapters = []
    books: dict[str, dict[str, object]] = {}
    book_order: list[str] = []

    for row in quiz_sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        chapter_id = int(row[0])
        book = value(row, 1)
        chapter = int(row[2])
        testament = "old" if book in OLD_TESTAMENT else "new"
        book_code = BOOK_CODES.get(book)
        if not book_code:
            raise ValueError(f"No bskorea book code mapped for '{book}' (chapter id {chapter_id})")

        if book not in books:
            books[book] = {
                "name": book,
                "testament": testament,
                "startId": chapter_id,
                "chapters": 0,
            }
            book_order.append(book)

        books[book]["chapters"] = int(books[book]["chapters"]) + 1

        chapters.append(
            {
                "id": chapter_id,
                "book": book,
                "chapter": chapter,
                "testament": testament,
                "question": value(row, 3),
                "hint": value(row, 4),
                "options": [value(row, 5), value(row, 6), value(row, 7), value(row, 8)],
                "answer": "①②③④".find(value(row, 9)) + 1,
                "answerText": value(row, 10),
                "link": f"https://bible.bskorea.or.kr/bible/NKRV/{book_code}.{chapter}",
            }
        )

    payload = {
        "meta": {
            "title": "Verbum",
            "totalChapters": len(chapters),
            "totalBooks": len(book_order),
        },
        "books": [books[name] for name in book_order],
        "chapters": chapters,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        "window.BIBLE_APP_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT} with {len(chapters)} chapters.")


if __name__ == "__main__":
    main()
